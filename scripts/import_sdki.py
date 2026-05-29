"""
scripts/import_sdki.py v3.2
Import workbook SDKI ke PostgreSQL dalam 4 stage.

Perubahan v3.2:
- Kolom bobot (lama, statis 1/2) dihapus dari semua tabel relasi
- Kolom bobot_khusus (NUMERIC) dibaca dari Excel dan diimport langsung
- Penyebab tidak lagi punya kolom bobot (tidak masuk scoring AI)
"""
import argparse, os, sys
import psycopg2
from openpyxl import load_workbook

def get_db(url):
    return psycopg2.connect(url, connect_timeout=10)

def norm_text(t):
    import re, unicodedata
    if not t: return ""
    t = "".join(c for c in unicodedata.normalize("NFKD", str(t).lower())
                if not unicodedata.combining(c))
    t = re.sub(r"[^a-z0-9\s]", " ", t)
    return re.sub(r"\s+", " ", t).strip()

def to_float(v):
    """Konversi nilai Excel ke float, None jika kosong/invalid."""
    if v is None: return None
    try: return float(str(v).replace(",","."))
    except (ValueError, TypeError): return None

# ── Stage 1: Diagnosa ─────────────────────────────────────────────
def stage1_diagnosa(wb, conn, truncate):
    if truncate:
        with conn.cursor() as cur:
            cur.execute("TRUNCATE diagnosa CASCADE")
        conn.commit()
    with conn.cursor() as cur:
        for row in wb["Diagnosa"].iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            kode_ref = row[0]; kode = row[1]; nama = row[2]
            kategori = row[3] if len(row) > 3 else None
            subkategori = row[4] if len(row) > 4 else None
            cur.execute("""
                INSERT INTO diagnosa (kode_ref,kode,nama,kategori,subkategori,is_aktif)
                VALUES (%s,%s,%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO UPDATE SET
                    kode=EXCLUDED.kode, nama=EXCLUDED.nama,
                    kategori=EXCLUDED.kategori, subkategori=EXCLUDED.subkategori
            """, (kode_ref, kode, nama, kategori, subkategori))
    conn.commit()
    with conn.cursor() as cur:
        cur.execute("SELECT kode_ref, id FROM diagnosa")
        return {r[0]: r[1] for r in cur.fetchall()}

# ── Stage 2: Gejala ───────────────────────────────────────────────
def stage2_gejala(wb, conn, diag_map):
    with conn.cursor() as cur: cur.execute("TRUNCATE gejala CASCADE")
    conn.commit()
    gejala_map = {}
    with conn.cursor() as cur:
        for row in wb["Gejala"].iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            gid_ref = row[0]; nama = row[1]; tipe = row[3]
            n = norm_text(nama)
            cur.execute("""
                INSERT INTO gejala (kode_ref,nama,norm,tipe,is_aktif)
                VALUES (%s,%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO UPDATE SET nama=EXCLUDED.nama, norm=EXCLUDED.norm
                RETURNING id
            """, (gid_ref, nama, n, tipe))
            gejala_map[gid_ref] = cur.fetchone()[0]
    conn.commit()

    # Alias gejala
    with conn.cursor() as cur: cur.execute("TRUNCATE gejala_alias CASCADE")
    conn.commit()
    with conn.cursor() as cur:
        for row in wb["Gejala_Alias"].iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]: continue
            aid = row[0]; gid_ref = row[1]; text = row[2]
            atype = row[3]; prio = row[4]
            db_gid = gejala_map.get(gid_ref)
            if not db_gid or not text: continue
            n = norm_text(text)
            cur.execute("""
                INSERT INTO gejala_alias
                  (kode_ref,id_gejala,alias_text,alias_norm,alias_type,priority,is_aktif)
                VALUES (%s,%s,%s,%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO NOTHING
            """, (aid, db_gid, text, n, atype or "variant", int(prio or 5)))
    conn.commit()

    # Relasi diagnosa-gejala — baca bobot_khusus dari kolom ke-6
    with conn.cursor() as cur: cur.execute("TRUNCATE diagnosa_gejala CASCADE")
    conn.commit()
    with conn.cursor() as cur:
        for row in wb["Diagnosa_Gejala"].iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1] or not row[2]: continue
            kref     = row[0]
            did_ref  = row[1]
            gid_ref  = row[2]
            kel      = row[3]
            tipe     = row[4]
            bk       = to_float(row[5])   # bobot_khusus — bisa None
            db_did   = diag_map.get(did_ref)
            db_gid   = gejala_map.get(gid_ref)
            if not db_did or not db_gid: continue
            cur.execute("""
                INSERT INTO diagnosa_gejala
                  (kode_ref,id_diagnosa,id_gejala,kelompok_gejala,tipe_gejala,
                   bobot_khusus,is_aktif)
                VALUES (%s,%s,%s,%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO NOTHING
            """, (kref, db_did, db_gid,
                  kel or "minor", tipe,
                  bk))
    conn.commit()
    print(f"  Gejala: {len(gejala_map)}")

# ── Stage 3: Penyebab ─────────────────────────────────────────────
def stage3_penyebab(wb, conn, diag_map):
    with conn.cursor() as cur: cur.execute("TRUNCATE penyebab CASCADE")
    conn.commit()
    p_map = {}
    with conn.cursor() as cur:
        for row in wb["Penyebab"].iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            pref = row[0]; nama = row[1]
            cur.execute("""
                INSERT INTO penyebab (kode_ref,nama,norm,is_aktif)
                VALUES (%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO UPDATE SET nama=EXCLUDED.nama, norm=EXCLUDED.norm
                RETURNING id
            """, (pref, nama, norm_text(nama)))
            p_map[pref] = cur.fetchone()[0]
    conn.commit()

    with conn.cursor() as cur: cur.execute("TRUNCATE penyebab_alias CASCADE")
    conn.commit()
    with conn.cursor() as cur:
        for row in wb["Penyebab_Alias"].iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]: continue
            aid = row[0]; pref = row[1]; text = row[2]
            atype = row[3]; prio = row[4]
            db_pid = p_map.get(pref)
            if not db_pid or not text: continue
            cur.execute("""
                INSERT INTO penyebab_alias
                  (kode_ref,id_penyebab,alias_text,alias_norm,alias_type,priority,is_aktif)
                VALUES (%s,%s,%s,%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO NOTHING
            """, (aid, db_pid, text, norm_text(text), atype or "variant", int(prio or 5)))
    conn.commit()

    with conn.cursor() as cur: cur.execute("TRUNCATE diagnosa_penyebab CASCADE")
    conn.commit()
    with conn.cursor() as cur:
        for row in wb["Diagnosa_Penyebab"].iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1] or not row[2]: continue
            kref = row[0]; did_ref = row[1]; pid_ref = row[2]
            db_did = diag_map.get(did_ref); db_pid = p_map.get(pid_ref)
            if not db_did or not db_pid: continue
            cur.execute("""
                INSERT INTO diagnosa_penyebab (kode_ref,id_diagnosa,id_penyebab,is_aktif)
                VALUES (%s,%s,%s,TRUE) ON CONFLICT (kode_ref) DO NOTHING
            """, (kref, db_did, db_pid))
    conn.commit()
    print(f"  Penyebab: {len(p_map)}")

# ── Stage 4: Faktor Risiko ────────────────────────────────────────
def stage4_risiko(wb, conn, diag_map):
    with conn.cursor() as cur: cur.execute("TRUNCATE faktor_risiko CASCADE")
    conn.commit()
    r_map = {}
    with conn.cursor() as cur:
        for row in wb["Faktor_Risiko"].iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            rref = row[0]; nama = row[1]
            cur.execute("""
                INSERT INTO faktor_risiko (kode_ref,nama,norm,is_aktif)
                VALUES (%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO UPDATE SET nama=EXCLUDED.nama, norm=EXCLUDED.norm
                RETURNING id
            """, (rref, nama, norm_text(nama)))
            r_map[rref] = cur.fetchone()[0]
    conn.commit()

    with conn.cursor() as cur: cur.execute("TRUNCATE faktor_risiko_alias CASCADE")
    conn.commit()
    with conn.cursor() as cur:
        for row in wb["Risiko_Alias"].iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]: continue
            aid = row[0]; rref = row[1]; text = row[2]
            atype = row[3]; prio = row[4]
            db_rid = r_map.get(rref)
            if not db_rid or not text: continue
            cur.execute("""
                INSERT INTO faktor_risiko_alias
                  (kode_ref,id_risiko,alias_text,alias_norm,alias_type,priority,is_aktif)
                VALUES (%s,%s,%s,%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO NOTHING
            """, (aid, db_rid, text, norm_text(text), atype or "variant", int(prio or 5)))
    conn.commit()

    # Relasi diagnosa-risiko — baca bobot_khusus dari kolom ke-4
    with conn.cursor() as cur: cur.execute("TRUNCATE diagnosa_risiko CASCADE")
    conn.commit()
    with conn.cursor() as cur:
        for row in wb["Diagnosa_Risiko"].iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1] or not row[2]: continue
            kref    = row[0]
            did_ref = row[1]
            rid_ref = row[2]
            bk      = to_float(row[3])  # bobot_khusus — bisa None
            db_did  = diag_map.get(did_ref)
            db_rid  = r_map.get(rid_ref)
            if not db_did or not db_rid: continue
            cur.execute("""
                INSERT INTO diagnosa_risiko
                  (kode_ref,id_diagnosa,id_risiko,bobot_khusus,is_aktif)
                VALUES (%s,%s,%s,%s,TRUE)
                ON CONFLICT (kode_ref) DO NOTHING
            """, (kref, db_did, db_rid, bk))
    conn.commit()
    print(f"  Faktor Risiko: {len(r_map)}")

# ── Main ──────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="Import workbook SDKI ke PostgreSQL (v3.2)"
    )
    ap.add_argument("workbook", help="Path ke file .xlsx")
    ap.add_argument("--db",
        default=os.getenv("DATABASE_URL",
                          "postgresql://postgres:admin@localhost:5432/rshu_ai"))
    ap.add_argument("--truncate", action="store_true",
                    help="Hapus semua data sebelum import")
    ap.add_argument("--dry-run", action="store_true",
                    help="Validasi tanpa menulis ke DB")
    args = ap.parse_args()

    print(f"Import: {args.workbook}")
    if args.dry_run:
        print("DRY RUN — tidak ada yang ditulis ke DB")
        wb = load_workbook(args.workbook, read_only=True, data_only=True)
        for sname in ["Diagnosa","Gejala","Gejala_Alias","Diagnosa_Gejala",
                      "Penyebab","Penyebab_Alias","Diagnosa_Penyebab",
                      "Faktor_Risiko","Risiko_Alias","Diagnosa_Risiko"]:
            ws = wb[sname]
            n = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[0])
            print(f"  {sname}: {n} baris")
        return

    wb   = load_workbook(args.workbook, read_only=True, data_only=True)
    conn = get_db(args.db)

    print("Stage 1 — Diagnosa...")
    diag_map = stage1_diagnosa(wb, conn, args.truncate)
    print(f"  Diagnosa: {len(diag_map)}")

    print("Stage 2 — Gejala & Diagnosa_Gejala (dengan bobot_khusus)...")
    stage2_gejala(wb, conn, diag_map)

    print("Stage 3 — Penyebab...")
    stage3_penyebab(wb, conn, diag_map)

    print("Stage 4 — Faktor Risiko & Diagnosa_Risiko (dengan bobot_khusus)...")
    stage4_risiko(wb, conn, diag_map)

    conn.close()
    print("\nImport selesai ✓")
    print("Catatan: bobot_khusus dibaca langsung dari Excel.")
    print("  NULL = engine gunakan bobot fallback otomatis (Mayor=3.0, Minor=1.0, Risiko=0.6)")
    print("  Isi angka = engine gunakan nilai tersebut (perawat bisa ubah di Excel)")

if __name__ == "__main__":
    main()
